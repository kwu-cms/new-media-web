#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excelファイルのプレゼンパスとレポートパスを学籍番号に基づいて自動設定
"""

import sys
from pathlib import Path

# パス設定
EXCEL_FILE = Path("data/students.xlsx")
REPORTS_DIR = Path("assets/reports")
PRESENTATIONS_DIR = Path("assets/presentations")

def update_excel_paths():
    """Excelファイルのパスを更新"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Excelファイルを読み込み
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # ヘッダー行を探す
        header_row = 1
        headers = {}
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        
        # プレゼンパス列のインデックスを取得
        presentation_col = None
        for col_name in ['プレゼンパス', 'プレゼン', 'プレゼンテーション', 'presentation']:
            if col_name in headers:
                presentation_col = headers[col_name]
                break
        
        # レポートパス列のインデックスを取得
        report_col = None
        for col_name in ['レポートパス', 'レポート', 'report']:
            if col_name in headers:
                report_col = headers[col_name]
                break
        
        # 学籍番号列のインデックスを取得
        student_id_col = None
        for col_name in ['学籍番号', 'student_id', 'ID']:
            if col_name in headers:
                student_id_col = headers[col_name]
                break
        
        if not student_id_col:
            print("エラー: 学籍番号列が見つかりません")
            return False
        
        print(f"学籍番号列: 列{student_id_col}")
        if presentation_col:
            print(f"プレゼンパス列: 列{presentation_col}")
        if report_col:
            print(f"レポートパス列: 列{report_col}\n")
        
        updated_presentation = 0
        updated_report = 0
        
        # 2行目以降を処理
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # 学籍番号を取得
            student_id_cell = row[student_id_col - 1]
            if not student_id_cell.value:
                continue
            
            student_id = str(student_id_cell.value).strip()
            
            # 7桁の数字か確認
            if not student_id.isdigit() or len(student_id) != 7:
                continue
            
            # プレゼンパスを設定
            if presentation_col:
                presentation_cell = row[presentation_col - 1]
                if not presentation_cell.value or str(presentation_cell.value).strip() == '':
                    presentation_path = f"{student_id}.pdf"
                    presentation_cell.value = presentation_path
                    updated_presentation += 1
                    print(f"✓ {student_id}: プレゼンパスを設定 → {presentation_path}")
            
            # レポートパスを設定
            if report_col:
                report_cell = row[report_col - 1]
                if not report_cell.value or str(report_cell.value).strip() == '':
                    report_path = f"{student_id}.pdf"
                    report_cell.value = report_path
                    updated_report += 1
                    print(f"✓ {student_id}: レポートパスを設定 → {report_path}")
        
        # Excelファイルを保存
        wb.save(EXCEL_FILE)
        print(f"\n{'='*50}")
        print(f"更新完了:")
        print(f"  プレゼンパス: {updated_presentation}名")
        print(f"  レポートパス: {updated_report}名")
        print(f"{'='*50}\n")
        
        return True
        
    except ImportError:
        print("エラー: openpyxlライブラリが必要です")
        print("インストール方法: pip3 install openpyxl")
        return False
    except Exception as e:
        print(f"エラー: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    if not EXCEL_FILE.exists():
        print(f"エラー: {EXCEL_FILE} が見つかりません")
        sys.exit(1)
    
    print("="*60)
    print("Excelファイルのパスを自動設定")
    print("="*60)
    print(f"\nExcelファイル: {EXCEL_FILE}\n")
    
    if update_excel_paths():
        print("✓ Excelファイルの更新が完了しました")
    else:
        print("✗ Excelファイルの更新に失敗しました")
        sys.exit(1)

if __name__ == "__main__":
    main()
