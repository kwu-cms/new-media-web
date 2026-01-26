#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
本番プレゼンテーションファイル処理スクリプト
PPTXファイルをPDFに変換し、学籍番号でリネームして配置、Excelを更新
"""

import os
import re
import subprocess
import sys
from pathlib import Path
import shutil

# パス設定（プロジェクトルートからの相対パス）
SCRIPT_DIR = Path(__file__).parent.parent.parent
INPUT_DIR = SCRIPT_DIR / "assets" / "presentations" / "final_input"  # 本番PPTXファイルを配置するフォルダ
OUTPUT_DIR = SCRIPT_DIR / "assets" / "presentations" / "rehearsal" / "pdf"  # PDF出力先
EXCEL_FILE = SCRIPT_DIR / "data" / "students.xlsx"

def check_libreoffice():
    """LibreOfficeが利用可能か確認"""
    for cmd in ['libreoffice', 'soffice']:
        try:
            result = subprocess.run(['which', cmd], capture_output=True, text=True)
            if result.returncode == 0:
                return result.stdout.strip()
        except:
            pass
    return None

def extract_student_id_from_filename(filename):
    """ファイル名から学籍番号を抽出（7桁の数字）"""
    # ファイル名から7桁の数字を探す
    match = re.search(r'\d{7}', filename)
    if match:
        return match.group()
    return None

def convert_pptx_to_pdf_libreoffice(pptx_path, output_dir):
    """LibreOfficeを使用してPPTXをPDFに変換"""
    libreoffice_path = check_libreoffice()
    if not libreoffice_path:
        print("エラー: LibreOfficeが見つかりません。")
        print("macOSの場合: brew install --cask libreoffice")
        return False
    
    try:
        # 出力ディレクトリが存在しない場合は作成
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # LibreOfficeでPDFに変換
        cmd = [
            libreoffice_path,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', str(output_dir),
            str(pptx_path)
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
        
        if result.returncode == 0:
            return True
        else:
            print(f"変換エラー: {result.stderr}")
            return False
    except subprocess.TimeoutExpired:
        print(f"タイムアウト: {pptx_path} の変換に時間がかかりすぎています")
        return False
    except Exception as e:
        print(f"変換エラー: {e}")
        return False

def rename_pdf_by_student_id(pdf_path, student_id):
    """PDFファイルを学籍番号でリネーム"""
    if not pdf_path.exists():
        return None
    
    new_path = pdf_path.parent / f"{student_id}.pdf"
    
    # 既存のファイルがある場合はバックアップ
    if new_path.exists():
        backup_path = new_path.parent / f"{student_id}.pdf.backup"
        shutil.move(str(new_path), str(backup_path))
        print(f"  ⚠️  既存のPDFをバックアップ: {backup_path.name}")
    
    try:
        shutil.move(str(pdf_path), str(new_path))
        return new_path
    except Exception as e:
        print(f"  ✗ リネームエラー: {e}")
        return None

def update_excel_presentation_path(student_id):
    """Excelファイルのプレゼンパスを更新"""
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # ヘッダー行を探す
        header_row = 1
        headers = {}
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        
        # プレゼンパス列と学籍番号列を取得
        presentation_col = None
        for col_name in ['プレゼンパス', 'プレゼン', 'プレゼンテーション']:
            if col_name in headers:
                presentation_col = headers[col_name]
                break
        
        student_id_col = None
        for col_name in ['学籍番号', 'student_id', 'ID']:
            if col_name in headers:
                student_id_col = headers[col_name]
                break
        
        if not student_id_col:
            print("  ⚠️  学籍番号列が見つかりません")
            return False
        
        # 該当する学生の行を探して更新
        updated = False
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            student_id_cell = row[student_id_col - 1]
            if not student_id_cell.value:
                continue
            
            if str(student_id_cell.value).strip() == student_id:
                if presentation_col:
                    presentation_cell = row[presentation_col - 1]
                    presentation_cell.value = f"{student_id}.pdf"
                    updated = True
                    print(f"  ✓ Excel更新: {student_id}.pdf")
                break
        
        if updated:
            wb.save(EXCEL_FILE)
        
        return updated
    except Exception as e:
        print(f"  ✗ Excel更新エラー: {e}")
        return False

def process_final_presentations():
    """本番PPTXファイルを処理"""
    print("="*60)
    print("本番プレゼンテーションファイル処理")
    print("="*60)
    print()
    
    # 入力ディレクトリの確認
    if not INPUT_DIR.exists():
        print(f"入力ディレクトリを作成します: {INPUT_DIR}")
        INPUT_DIR.mkdir(parents=True, exist_ok=True)
        print(f"\n以下のフォルダに本番PPTXファイルを配置してください:")
        print(f"  {INPUT_DIR}")
        print("\nファイルを配置したら、このスクリプトを再実行してください。")
        return
    
    # PPTXファイルを検索
    pptx_files = list(INPUT_DIR.glob("*.pptx")) + list(INPUT_DIR.glob("*.ppt"))
    
    if not pptx_files:
        print(f"PPTXファイルが見つかりません: {INPUT_DIR}")
        print("\n以下のフォルダに本番PPTXファイルを配置してください:")
        print(f"  {INPUT_DIR}")
        return
    
    print(f"見つかったPPTXファイル: {len(pptx_files)}個\n")
    
    # 出力ディレクトリの確認
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    success_count = 0
    failed_count = 0
    processed_students = []
    
    for pptx_file in pptx_files:
        print(f"\n処理中: {pptx_file.name}")
        print("-" * 60)
        
        # 学籍番号を抽出
        student_id = extract_student_id_from_filename(pptx_file.name)
        
        if not student_id:
            print(f"  ✗ 学籍番号が見つかりません: {pptx_file.name}")
            print(f"    ファイル名に7桁の数字（学籍番号）が含まれている必要があります")
            failed_count += 1
            continue
        
        print(f"  ✓ 学籍番号を検出: {student_id}")
        
        # PDFに変換
        print(f"  → PDFに変換中...")
        if not convert_pptx_to_pdf_libreoffice(pptx_file, OUTPUT_DIR):
            print(f"  ✗ PDF変換に失敗しました")
            failed_count += 1
            continue
        
        # 変換されたPDFファイルを探す（元のファイル名ベース）
        pdf_name = pptx_file.stem + ".pdf"
        pdf_path = OUTPUT_DIR / pdf_name
        
        if not pdf_path.exists():
            print(f"  ✗ 変換されたPDFが見つかりません: {pdf_name}")
            failed_count += 1
            continue
        
        print(f"  ✓ PDF変換完了: {pdf_path.name}")
        
        # 学籍番号でリネーム
        print(f"  → リネーム中...")
        renamed_path = rename_pdf_by_student_id(pdf_path, student_id)
        
        if not renamed_path:
            print(f"  ✗ リネームに失敗しました")
            failed_count += 1
            continue
        
        print(f"  ✓ リネーム完了: {renamed_path.name}")
        
        # Excelを更新
        print(f"  → Excel更新中...")
        if update_excel_presentation_path(student_id):
            processed_students.append(student_id)
            success_count += 1
        else:
            print(f"  ⚠️  Excel更新をスキップしました（手動で更新してください）")
            success_count += 1  # PDF変換とリネームは成功している
    
    # 結果サマリー
    print("\n" + "="*60)
    print("処理完了")
    print("="*60)
    print(f"成功: {success_count}個")
    print(f"失敗: {failed_count}個")
    
    if processed_students:
        print(f"\n処理された学籍番号: {', '.join(processed_students)}")
        print(f"\n出力先: {OUTPUT_DIR}")
        print(f"\n次のステップ:")
        print(f"1. ウェブサイトをリロードして確認")
        print(f"2. 必要に応じてExcelファイルを手動で確認・修正")
    else:
        print("\n処理されたファイルがありません。")
    
    print("="*60)

if __name__ == "__main__":
    try:
        process_final_presentations()
    except KeyboardInterrupt:
        print("\n\n処理が中断されました。")
        sys.exit(1)
    except Exception as e:
        print(f"\n予期しないエラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
