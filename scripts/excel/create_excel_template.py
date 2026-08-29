#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel雛形ファイル作成スクリプト
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# パス設定（プロジェクトルートからの相対パス）
SCRIPT_DIR = Path(__file__).parent.parent.parent
EXCEL_FILE = SCRIPT_DIR / "data" / "students.xlsx"

# ワークブックを作成
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "学生データ"

# ヘッダー行を定義
headers = [
    "No",
    "所属学年",
    "学籍番号",
    "氏名",
    "氏名英字",
    "題目",
    "画像パス",
    "レポートパス",
    "プレゼンパス"
]

# サンプルデータ（個人情報は含めない）
sample_data = [
    [1, "メデ4", "1234567", "（非公開）", "（Non-public）", "研究題目の例", "image1.png", "1234567.pdf", "1234567.pdf"],
]

# ヘッダー行を書き込み
header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 列幅を調整
column_widths = {
    "A": 8,   # No
    "B": 12,  # 所属学年
    "C": 15,  # 学籍番号
    "D": 15,  # 氏名
    "E": 20,  # 氏名英字
    "F": 40,  # 題目
    "G": 30,  # 画像パス
    "H": 30,  # レポートパス
    "I": 30,  # プレゼンパス
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# サンプルデータを書き込み
for row_num, row_data in enumerate(sample_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col_num == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[1].height = 25

# 説明シートを追加
ws2 = wb.create_sheet("説明")
ws2.title = "説明"

instructions = [
    ["Excelファイル項目説明", "", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", ""],
    ["列名", "説明", "例", "", "", "", "", "", ""],
    ["No", "学生番号（連番）", "1, 2, 3...", "", "", "", "", "", ""],
    ["所属学年", "所属学年", "メデ4", "", "", "", "", "", ""],
    ["学籍番号", "学生の学籍番号", "1234567", "", "", "", "", "", ""],
    ["氏名", "学生の氏名（日本語）", "（非公開）", "", "", "", "", "", ""],
    ["氏名英字", "学生の氏名（ローマ字）", "（Non-public）", "", "", "", "", "", ""],
    ["題目", "卒業研究の題目", "研究題目の例", "", "", "", "", "", ""],
    ["画像パス", "画像ファイル名（複数可、カンマ区切り）", "image1.png", "", "", "", "", "", ""],
    ["レポートパス", "Wordレポートファイル名（複数可、カンマ区切り）", "1234567.pdf", "", "", "", "", "", ""],
    ["プレゼンパス", "PPTXプレゼン資料ファイル名（複数可、カンマ区切り）", "1234567.pdf", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", "", ""],
    ["注意事項", "", "", "", "", "", "", "", ""],
    ["1. ファイル名に日本語や特殊文字を含めないことを推奨します", "", "", "", "", "", "", "", ""],
    ["2. 画像ファイルは assets/images/ フォルダに配置してください", "", "", "", "", "", "", "", ""],
    ["3. レポートファイルは assets/reports/ フォルダに配置してください", "", "", "", "", "", "", "", ""],
    ["4. プレゼン資料ファイルは assets/presentations/ フォルダに配置してください", "", "", "", "", "", "", "", ""],
    ["5. 複数のファイルを指定する場合は、カンマ区切りで入力してください", "", "", "", "", "", "", "", ""],
]

for row_num, row_data in enumerate(instructions, 1):
    for col_num, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_num, column=col_num, value=value)
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
        elif row_num == 3:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 50
ws2.column_dimensions["C"].width = 50

wb.save(EXCEL_FILE)
print(f"Excel雛形ファイルを作成しました: {EXCEL_FILE}")
print(f"サンプル行: {len(sample_data)} 件")
