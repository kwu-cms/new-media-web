#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excelファイルの画像パス、レポートパス、プレゼンパスを自動設定するスクリプト
"""

import sys
from pathlib import Path
import re

# パス設定（プロジェクトルートからの相対パス）
SCRIPT_DIR = Path(__file__).parent.parent.parent
EXCEL_FILE = SCRIPT_DIR / "data" / "students.xlsx"
IMAGES_DIR = SCRIPT_DIR / "assets" / "images"
REPORTS_DIR = SCRIPT_DIR / "assets" / "reports" / "pdf"
PRESENTATIONS_DIR = SCRIPT_DIR / "assets" / "presentations" / "rehearsal" / "pdf"

# 画像ファイルの拡張子
IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.svg', '.PNG', '.JPG', '.JPEG', '.GIF', '.SVG', '.webp', '.WEBP'}

def get_image_files(student_id, max_count=10):
    """学籍番号フォルダ内の画像ファイルを取得（imageで始まるファイルを優先）"""
    student_folder = IMAGES_DIR / student_id
    
    if not student_folder.exists():
        return []
    
    # 画像ファイルを取得
    image_files = []
    for file_path in sorted(student_folder.iterdir()):
        if file_path.is_file() and file_path.suffix in IMAGE_EXTENSIONS:
            # imageで始まるファイルを優先
            priority = 0 if file_path.name.lower().startswith('image') else 1
            image_files.append((file_path.name, priority))
    
    # 優先度順（imageで始まるファイルが先）、その後ファイル名順にソート
    image_files.sort(key=lambda x: (x[1], x[0]))
    
    # ファイル名を返す（カンマ区切り用）
    selected = [name for name, _ in image_files[:max_count]]
    return selected

def update_excel_all_paths():
    """Excelファイルのすべてのパスを更新"""
    try:
        import openpyxl
        
        # Excelファイルを読み込み
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # ヘッダー行を探す
        header_row = 1
        headers = {}
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        
        # 各列のインデックスを取得
        image_col = None
        for col_name in ['画像パス', '画像', '画像ファイル', 'image']:
            if col_name in headers:
                image_col = headers[col_name]
                break
        
        report_col = None
        for col_name in ['レポートパス', 'レポート', 'report']:
            if col_name in headers:
                report_col = headers[col_name]
                break
        
        presentation_col = None
        for col_name in ['プレゼンパス', 'プレゼン', 'プレゼンテーション', 'presentation']:
            if col_name in headers:
                presentation_col = headers[col_name]
                break
        
        # 学籍番号列のインデックスを取得
        student_id_col = None
        for col_name in ['学籍番号', 'student_id', 'ID']:
            if col_name in headers:
                student_id_col = headers[col_name]
                break
        
        if not student_id_col:
            print("エラー: 学籍番号列が見つかりません")
            print(f"見つかった列: {list(headers.keys())}")
            return False
        
        print(f"学籍番号列: 列{student_id_col}")
        if image_col:
            print(f"画像パス列: 列{image_col}")
        if report_col:
            print(f"レポートパス列: 列{report_col}")
        if presentation_col:
            print(f"プレゼンパス列: 列{presentation_col}")
        print()
        
        updated_image = 0
        updated_report = 0
        updated_presentation = 0
        
        # 2行目以降を処理
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # 学籍番号を取得
            student_id_cell = row[student_id_col - 1]
            if not student_id_cell.value:
                continue
            
            student_id = str(student_id_cell.value).strip()
            
            # 7桁の数字か確認
            if not re.match(r'^\d{7}$', student_id):
                continue
            
            # 画像パスを設定
            if image_col:
                image_cell = row[image_col - 1]
                if not image_cell.value or str(image_cell.value).strip() == '':
                    image_files = get_image_files(student_id)
                    if image_files:
                        image_paths = ','.join(image_files)
                        image_cell.value = image_paths
                        updated_image += 1
                        print(f"✓ {student_id}: 画像パスを設定 ({len(image_files)}個)")
                    else:
                        print(f"⚠️  {student_id}: 画像ファイルが見つかりません")
            
            # レポートパスを設定
            if report_col:
                report_cell = row[report_col - 1]
                if not report_cell.value or str(report_cell.value).strip() == '':
                    report_path = f"{student_id}.pdf"
                    report_file = REPORTS_DIR / report_path
                    if report_file.exists():
                        report_cell.value = report_path
                        updated_report += 1
                        print(f"✓ {student_id}: レポートパスを設定 → {report_path}")
                    else:
                        print(f"⚠️  {student_id}: レポートファイルが見つかりません ({report_path})")
            
            # プレゼンパスを設定
            if presentation_col:
                presentation_cell = row[presentation_col - 1]
                if not presentation_cell.value or str(presentation_cell.value).strip() == '':
                    presentation_path = f"{student_id}.pdf"
                    presentation_file = PRESENTATIONS_DIR / presentation_path
                    if presentation_file.exists():
                        presentation_cell.value = presentation_path
                        updated_presentation += 1
                        print(f"✓ {student_id}: プレゼンパスを設定 → {presentation_path}")
                    else:
                        print(f"⚠️  {student_id}: プレゼンファイルが見つかりません ({presentation_path})")
        
        # Excelファイルを保存
        wb.save(EXCEL_FILE)
        print(f"\n{'='*60}")
        print(f"更新完了:")
        print(f"  画像パス: {updated_image}名")
        print(f"  レポートパス: {updated_report}名")
        print(f"  プレゼンパス: {updated_presentation}名")
        print(f"{'='*60}\n")
        
        return True
        
    except ImportError:
        print("エラー: openpyxlライブラリが必要です")
        print("インストール方法: pip3 install openpyxl")
        return False
    except Exception as e:
        print(f"エラー: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    # ディレクトリの確認
    if not EXCEL_FILE.exists():
        print(f"エラー: {EXCEL_FILE} が見つかりません")
        sys.exit(1)
    
    print("="*60)
    print("Excelファイルのパスを自動設定（画像・レポート・プレゼン）")
    print("="*60)
    print(f"\nExcelファイル: {EXCEL_FILE}")
    print(f"画像フォルダ: {IMAGES_DIR}")
    print(f"レポートフォルダ: {REPORTS_DIR}")
    print(f"プレゼンフォルダ: {PRESENTATIONS_DIR}\n")
    
    if update_excel_all_paths():
        print("✓ Excelファイルの更新が完了しました")
    else:
        print("✗ Excelファイルの更新に失敗しました")
        sys.exit(1)

if __name__ == "__main__":
    main()
