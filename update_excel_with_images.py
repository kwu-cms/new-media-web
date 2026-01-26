#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excelファイルの画像パス列を更新するスクリプト
各学籍番号フォルダ内の大きな画像ファイルを選択してExcelに反映
"""

import sys
from pathlib import Path
import re

# パス設定
ZIP_DIR = Path("assets/images")  # zipフォルダはimagesにリネーム済み
EXCEL_FILE = Path("data/students.xlsx")

# 画像ファイルの拡張子
IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.svg', '.PNG', '.JPG', '.JPEG', '.GIF', '.SVG', '.webp', '.WEBP'}

def get_large_images(student_id, max_count=5):
    """学籍番号フォルダ内の大きな画像ファイルを取得"""
    student_folder = ZIP_DIR / student_id
    
    if not student_folder.exists():
        return []
    
    # 画像ファイルをファイルサイズ順にソート
    image_files = []
    for file_path in student_folder.iterdir():
        if file_path.is_file() and file_path.suffix in IMAGE_EXTENSIONS:
            try:
                size = file_path.stat().st_size
                image_files.append((file_path.name, size))
            except:
                pass
    
    # ファイルサイズの大きい順にソート
    image_files.sort(key=lambda x: x[1], reverse=True)
    
    # 上位のファイル名を返す（カンマ区切り）
    selected = [name for name, _ in image_files[:max_count]]
    return selected

def update_excel():
    """Excelファイルを更新"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Excelファイルを読み込み
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # ヘッダー行を探す
        header_row = 1
        headers = {}
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        
        # 画像パス列のインデックスを取得（複数の可能性に対応）
        image_col = None
        for col_name in ['画像パス', '画像', '画像ファイル', 'image']:
            if col_name in headers:
                image_col = headers[col_name]
                break
        
        if not image_col:
            print("エラー: 画像パス列が見つかりません")
            print(f"見つかった列: {list(headers.keys())}")
            return False
        
        print(f"画像パス列: 列{image_col} ({list(headers.keys())[list(headers.values()).index(image_col)]})")
        
        # 学籍番号列のインデックスを取得
        student_id_col = None
        for col_name in ['学籍番号', '学籍番号', 'student_id', 'ID']:
            if col_name in headers:
                student_id_col = headers[col_name]
                break
        
        if not student_id_col:
            print("エラー: 学籍番号列が見つかりません")
            return False
        
        print(f"学籍番号列: 列{student_id_col}\n")
        
        updated_count = 0
        
        # 2行目以降を処理
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # 学籍番号を取得
            student_id_cell = row[student_id_col - 1]
            if not student_id_cell.value:
                continue
            
            student_id = str(student_id_cell.value).strip()
            
            # 7桁の数字か確認
            if not re.match(r'^\d{7}$', student_id):
                continue
            
            # 大きな画像ファイルを取得
            large_images = get_large_images(student_id)
            
            if not large_images:
                print(f"⚠️  {student_id}: 画像ファイルが見つかりません")
                continue
            
            # 画像パスをカンマ区切りで設定
            image_paths = ','.join(large_images)
            
            # Excelのセルを更新
            image_cell = row[image_col - 1]
            old_value = image_cell.value if image_cell.value else ""
            image_cell.value = image_paths
            
            print(f"✓ {student_id}: {len(large_images)}個の画像を設定")
            if old_value:
                print(f"  旧値: {old_value}")
            print(f"  新値: {image_paths[:80]}...")
            
            updated_count += 1
        
        # Excelファイルを保存
        wb.save(EXCEL_FILE)
        print(f"\n{'='*50}")
        print(f"更新完了: {updated_count}名の学生データを更新しました")
        print(f"{'='*50}\n")
        
        return True
        
    except ImportError:
        print("エラー: openpyxlライブラリが必要です")
        print("インストール方法: pip3 install openpyxl")
        return False
    except Exception as e:
        print(f"エラー: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    # ディレクトリの確認
    if not ZIP_DIR.exists():
        print(f"エラー: {ZIP_DIR} が見つかりません")
        sys.exit(1)
    
    if not EXCEL_FILE.exists():
        print(f"エラー: {EXCEL_FILE} が見つかりません")
        sys.exit(1)
    
    print("="*60)
    print("Excelファイルの画像パス列を更新")
    print("="*60)
    print(f"\n画像フォルダ: {ZIP_DIR}")
    print(f"Excelファイル: {EXCEL_FILE}\n")
    
    # 各学籍番号フォルダの画像ファイルを確認
    student_folders = sorted([d for d in ZIP_DIR.iterdir() if d.is_dir() and re.match(r'^\d{7}$', d.name)])
    
    if not student_folders:
        print("学籍番号フォルダが見つかりません")
        sys.exit(1)
    
    print(f"見つかった学籍番号フォルダ: {len(student_folders)}個\n")
    
    # 各フォルダの画像ファイルを確認
    for student_folder in student_folders:
        student_id = student_folder.name
        images = get_large_images(student_id, max_count=3)
        if images:
            print(f"{student_id}: {len(images)}個の大きな画像ファイルを検出")
            for img in images[:3]:
                img_path = student_folder / img
                size_mb = img_path.stat().st_size / (1024 * 1024)
                print(f"  - {img} ({size_mb:.2f} MB)")
        else:
            print(f"{student_id}: 画像ファイルが見つかりません")
    
    print(f"\n{'='*60}\n")
    
    # Excelを更新
    if update_excel():
        print("✓ Excelファイルの更新が完了しました")
    else:
        print("✗ Excelファイルの更新に失敗しました")
        sys.exit(1)

if __name__ == "__main__":
    main()
